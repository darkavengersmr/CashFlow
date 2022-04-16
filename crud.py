import os
from datetime import datetime, timedelta

from database import database
from sqlalchemy import and_, func
from models import users, inflows, inflows_regular, outflows, outflows_regular, assets, liabilities, categories
import schemas

from openpyxl import load_workbook
from openpyxl.styles import Font

async def get_user(user_id: int):
    result = await database.fetch_one(users.select().where(users.c.id == user_id))
    if result:
        return dict(result)
    else:
        return None


async def get_user_by_email(email: str):
    return await database.fetch_one(users.select().where(users.c.email == email))


async def get_user_by_username(username: str):
    return await database.fetch_one(users.select().where(users.c.username == username))


async def get_users(skip: int = 0, limit: int = 100):
     results = await database.fetch_all(users.select().offset(skip).limit(limit))
     return [dict(result) for result in results]


async def create_user(user: schemas.UserCreate, hashed_password: str):
    db_user = users.insert().values(username=user.username, email=user.email, hashed_password=hashed_password, is_active=user.is_active)
    user_id = await database.execute(db_user)
    return schemas.User(**user.dict(), id=user_id)


async def create_user_inflow(inflow: schemas.InflowCreate, user_id: int):
    query = inflows.insert().values(**inflow.dict(), owner_id=user_id)
    inflow_id = await database.execute(query)
    return schemas.InflowInDB(**inflow.dict(), id=inflow_id, owner_id=user_id)


async def get_inflow_user(user_id: int, date_in: datetime, date_out: datetime):
    result = dict()
    list_inflows = await database.fetch_all(inflows.select().where(and_(inflows.c.owner_id == user_id,
                                                                        inflows.c.date >= date_in,
                                                                        inflows.c.date <= date_out)))
    result.update({"inflow": [dict(result) for result in list_inflows]})
    return result


async def delete_inflow_user(inflow_id: int, user_id: int):
    query = inflows.select().where(and_(inflows.c.id == inflow_id, inflows.c.owner_id == user_id))
    result = await database.execute(query)
    if result:
        query = inflows.delete().where(and_(inflows.c.id == inflow_id, inflows.c.owner_id == user_id))
        await database.execute(query)
        result = {"result": "inflow deleted"}
    else:
        result = {"result": "inflow for delete not found"}
    return result


async def create_user_inflow_regular(inflow_regular: schemas.InflowRegularCreate, user_id: int):
    query = inflows_regular.insert().values(**inflow_regular.dict(), owner_id=user_id)
    inflow_regular_id = await database.execute(query)
    return schemas.InflowRegularInDB(**inflow_regular.dict(), id=inflow_regular_id, owner_id=user_id)


async def get_inflow_regular_user(user_id: int):
    result = dict()
    query = "SELECT inflow_regular.description, inflow_lastsum.sum, inflow_regular.id, inflow_regular.owner_id " \
            "FROM inflow_regular, inflow_lastsum WHERE inflow_regular.owner_id = :owner_id " \
            "AND inflow_regular.owner_id = inflow_lastsum.owner_id " \
            "AND inflow_regular.description = inflow_lastsum.description " \
            "UNION SELECT description, sum, id, owner_id FROM inflow_regular WHERE description NOT IN (" \
            "SELECT description FROM inflow WHERE owner_id = :owner_id) AND owner_id = :owner_id"

    list_inflows_regular = await database.fetch_all(query=query, values={"owner_id": user_id})
    '''
    list_inflows_regular = await database.fetch_all(
        inflows_regular.select().where(inflows_regular.c.owner_id == user_id))
    '''
    result.update({"inflow_regular": [dict(result) for result in list_inflows_regular]})
    return result


async def update_user_inflow_regular(user_id: int, inflow_regular: schemas.InflowRegularOut):
    query = inflows_regular.select().where(and_(inflows_regular.c.id == inflow_regular.id,
                                           inflows_regular.c.owner_id == user_id))
    result = await database.execute(query)
    if result:
        query = inflows_regular.update().where(
            and_(inflows_regular.c.id == inflow_regular.id, inflows_regular.c.owner_id == user_id)).values(
            description=inflow_regular.description, sum=inflow_regular.sum)
        await database.execute(query)
        result = {"result": "inflow regular updated"}
    else:
        result = {"result": "inflow regular for update not found"}
    return result


async def delete_inflow_regular_user(inflow_regular_id: int, user_id: int):
    query = inflows_regular.select().where(and_(inflows_regular.c.id == inflow_regular_id,
                                                inflows_regular.c.owner_id == user_id))
    result = await database.execute(query)
    if result:
        query = inflows_regular.delete().where(and_(inflows_regular.c.id == inflow_regular_id,
                                               inflows_regular.c.owner_id == user_id))
        await database.execute(query)
        result = {"result": "regular inflow deleted"}
    else:
        result = {"result": "regular inflow for delete not found"}
    return result


async def create_user_outflow(outflow: schemas.OutflowCreate, user_id: int):
    query = outflows.insert().values(**outflow.dict(), owner_id=user_id)
    outflow_id = await database.execute(query)
    return schemas.OutflowInDB(**outflow.dict(), id=outflow_id, owner_id=user_id)


async def get_outflow_user(user_id: int, date_in: datetime, date_out: datetime):
    '''
    # вариант c группировкой и суммированием по аналогичным расходам за период
    result = dict()
    query = "SELECT description, sum(sum) as sum, count(description) as count FROM outflow " \
            "WHERE owner_id = :owner_id AND date >= :date_in AND date <= :date_out " \
            "GROUP BY description, owner_id " \
            "ORDER BY count DESC"
    list_outflows = await database.fetch_all(query=query, values={"owner_id": user_id, "date_in": date_in,
                                                                  "date_out": date_out})
    '''
    # вариант без группировки и суммирования по аналогичным расходам за период
    result = dict()
    list_outflows = await database.fetch_all(outflows.select().where(
        and_(outflows.c.owner_id == user_id, outflows.c.date >= date_in, outflows.c.date <= date_out)))
    result.update({"outflow": [dict(result) for result in list_outflows]})
    return result


async def delete_outflow_user(outflow_id: int, user_id: int):
    query = outflows.select().where(and_(outflows.c.id == outflow_id, outflows.c.owner_id == user_id))
    result = await database.execute(query)
    if result:
        query = outflows.delete().where(and_(outflows.c.id == outflow_id, outflows.c.owner_id == user_id))
        await database.execute(query)
        result = {"result": "outflow deleted"}
    else:
        result = {"result": "outflow for delete not found"}
    return result


async def create_user_outflow_regular(outflow_regular: schemas.OutflowRegularCreate, user_id: int):
    query = outflows_regular.insert().values(**outflow_regular.dict(), owner_id=user_id)
    outflow_regular_id = await database.execute(query)
    return schemas.OutflowRegularInDB(**outflow_regular.dict(), id=outflow_regular_id, owner_id=user_id)


async def get_outflow_regular_user(user_id: int):
    result = dict()
    query = "SELECT outflow_regular.description, outflow_lastsum.sum, outflow_regular.id, outflow_regular.owner_id " \
            "FROM outflow_regular, outflow_lastsum WHERE outflow_regular.owner_id = :owner_id " \
            "AND outflow_regular.owner_id = outflow_lastsum.owner_id " \
            "AND outflow_regular.description = outflow_lastsum.description " \
            "UNION SELECT description, sum, id, owner_id FROM outflow_regular WHERE description NOT IN (" \
            "SELECT description FROM outflow WHERE owner_id = :owner_id) AND owner_id = :owner_id"

    list_outflows_regular = await database.fetch_all(query=query, values={"owner_id": user_id})
    '''
    list_outflows_regular = await database.fetch_all(
        outflows_regular.select().where(outflows_regular.c.owner_id == user_id))
    '''
    result.update({"outflow_regular": [dict(result) for result in list_outflows_regular]})
    return result


async def update_user_outflow_regular(user_id: int, outflow_regular: schemas.OutflowRegularOut):
    query = outflows_regular.select().where(and_(outflows_regular.c.id == outflow_regular.id,
                                                 outflows_regular.c.owner_id == user_id))
    result = await database.execute(query)
    if result:
        query = outflows_regular.update().where(
            and_(outflows_regular.c.id == outflow_regular.id, outflows_regular.c.owner_id == user_id)).values(
            description=outflow_regular.description, sum=outflow_regular.sum)
        await database.execute(query)
        result = {"result": "outflow regular updated"}
    else:
        result = {"result": "outflow regular for update not found"}
    return result


async def delete_outflow_regular_user(outflow_regular_id: int, user_id: int):
    query = outflows_regular.select().where(and_(outflows_regular.c.id == outflow_regular_id,
                                                 outflows_regular.c.owner_id == user_id))
    result = await database.execute(query)
    if result:
        query = outflows_regular.delete().where(and_(outflows_regular.c.id == outflow_regular_id,
                                                outflows_regular.c.owner_id == user_id))
        await database.execute(query)
        result = {"result": "regular outflow deleted"}
    else:
        result = {"result": "regular outflow for delete not found"}
    return result


async def create_user_asset(asset: schemas.AssetCreate, user_id: int):
    query = assets.insert().values(**asset.dict(), owner_id=user_id)
    asset_id = await database.execute(query)
    return schemas.AssetInDB(**asset.dict(), id=asset_id, owner_id=user_id)


async def get_assets_user(user_id: int, date: datetime):
    result = dict()
    list_assets = await database.fetch_all(assets.select().where(and_(assets.c.owner_id == user_id,
                                                                 assets.c.date_in <= date,
                                                                 date <= assets.c.date_out)))
    result.update({"assets": [dict(result) for result in list_assets]})
    return result


async def update_user_asset(asset: schemas.AssetOut, user_id: int):
    query = assets.select().where(and_(assets.c.id == asset.id, assets.c.owner_id == user_id))
    result = await database.execute(query)
    if result:
        dt = asset.date
        dt_end = datetime.strptime(f"{dt.timetuple().tm_year}-{dt.timetuple().tm_mon}-01 00:00:00", "%Y-%m-%d %H:%M:%S") \
                 - timedelta(seconds=1)
        query = assets.update().where(and_(assets.c.id == asset.id, assets.c.owner_id == user_id)).values(
            date_out=dt_end)
        await database.execute(query)
        if hasattr(asset, 'sum') and asset.sum > 0:
            dt = datetime.now()
            dt_begin = datetime.strptime(f"{dt.timetuple().tm_year}-{dt.timetuple().tm_mon}-01 00:00:00",
                                         "%Y-%m-%d %H:%M:%S")
            query = assets.insert().values(date_in=dt_begin, date_out=datetime.now() + timedelta(days=100000),
                                           description=asset.description, sum=asset.sum, owner_id=user_id)
            await database.execute(query)
            result = {"result": "asset updated"}
        else:
            result = {"result": "asset deleted"}
    else:
        result = {"result": "asset for update/delete not found"}
    return result


async def create_user_liabilitie(liabilitie: schemas.LiabilitieCreate, user_id: int):
    query = liabilities.insert().values(**liabilitie.dict(), owner_id=user_id)
    liabilitie_id = await database.execute(query)
    return schemas.LiabilitieInDB(**liabilitie.dict(), id=liabilitie_id, owner_id=user_id)


async def get_liabilities_user(user_id: int, date: datetime):
    result = dict()
    list_liabilities = await database.fetch_all(liabilities.select().where(and_(liabilities.c.owner_id == user_id,
                                                                                liabilities.c.date_in <= date,
                                                                                date <= liabilities.c.date_out)))
    result.update({"liabilities": [dict(result) for result in list_liabilities]})
    return result


async def update_user_liabilitie(liabilitie: schemas.LiabilitieOut, user_id: int):
    query = liabilities.select().where(and_(liabilities.c.id == liabilitie.id, liabilities.c.owner_id == user_id))
    result = await database.execute(query)
    if result:
        dt = liabilitie.date
        dt_end = datetime.strptime(f"{dt.timetuple().tm_year}-{dt.timetuple().tm_mon}-01 00:00:00", "%Y-%m-%d %H:%M:%S")\
                 - timedelta(seconds=1)
        query = liabilities.update().where(
            and_(liabilities.c.id == liabilitie.id, liabilities.c.owner_id == user_id))\
            .values(date_out=dt_end)
        await database.execute(query)
        if hasattr(liabilitie, 'sum') and liabilitie.sum > 0:
            dt = datetime.now()
            dt_begin = datetime.strptime(f"{dt.timetuple().tm_year}-{dt.timetuple().tm_mon}-01 00:00:00",
                                         "%Y-%m-%d %H:%M:%S")
            query = liabilities.insert().values(date_in=dt_begin,
                                                date_out=datetime.now() + timedelta(days=100000),
                                                description=liabilitie.description,
                                                sum=liabilitie.sum,
                                                owner_id=user_id)
            await database.execute(query)
            result = {"result": "liabilitie updated"}
        else:
            result = {"result": "liabilitie deleted"}
    else:
        result = {"result": "liabilitie for update/delete not found"}
    return result


async def create_user_category(category: schemas.CategoryCreate, user_id: int):
    query = categories.insert().values(**category.dict(), owner_id=user_id)
    category_id = await database.execute(query)
    return schemas.CategoryInDB(**category.dict(), id=category_id, owner_id=user_id)


async def get_user_categories(user_id: int):
    result = dict()
    list_categories = await database.fetch_all(categories.select().where(categories.c.owner_id == user_id))
    result.update({"categories": [dict(result) for result in list_categories]})
    return result


async def delete_user_category(category_id: int, user_id: int):
    query = "SELECT category_id FROM assets WHERE owner_id = :owner_id AND category_id = :category_id"
    categories_in_asset = await database.fetch_all(query=query, values={"owner_id": user_id,
                                                                        "category_id": category_id})
    query = "SELECT category_id FROM liabilities WHERE owner_id = :owner_id AND category_id = :category_id"
    categories_in_liabilities = await database.fetch_all(query=query, values={"owner_id": user_id,
                                                                              "category_id": category_id})
    if categories_in_asset or categories_in_liabilities:
        result = {"result": "category in use, not deleted"}
    else:
        query = categories.select().where(and_(categories.c.id == category_id, categories.c.owner_id == user_id))
        result = await database.execute(query)
        if result:
            query = categories.delete().where(and_(categories.c.id == category_id, categories.c.owner_id == user_id))
            await database.execute(query)
            result = {"result": "category deleted"}
        else:
            result = {"result": "category for delete not found"}
    return result


async def get_assets_by_categories(user_id: int, date: datetime, asset: str):
    result = dict()
    query = "SELECT category_id, sum(sum) as sum FROM " + asset + " " \
            "WHERE owner_id = :owner_id AND date_in <= :date AND date_out >= :date " \
            "GROUP BY category_id, owner_id"
    list_categories = await database.fetch_all(query=query, values={"owner_id": user_id, "date": date})
    result.update({"categories": [dict(result) for result in list_categories]})
    return result


async def get_reports(user_id: int):
    out = dict()
    assets_result = []
    liabilities_result = []

    dt = datetime.now()
    while True:
        date = datetime.strptime(f"{dt.timetuple().tm_year}-{dt.timetuple().tm_mon}-15 00:00:00", "%Y-%m-%d %H:%M:%S")
        this_month_begin = datetime.strptime(f"{dt.timetuple().tm_year}-{dt.timetuple().tm_mon}-01 00:00:00",
                                             "%Y-%m-%d %H:%M:%S")

        list_assets = await database.fetch_all(assets.select().where(and_(assets.c.owner_id == user_id,
                                                                          assets.c.date_in <= date,
                                                                          date <= assets.c.date_out)))
        month_assets = [dict(result) for result in list_assets]
        month_assets_sum = 0
        for month_asset in month_assets:
            month_assets_sum += month_asset['sum']
        month_date = str(date)[5:7] + "." + str(date)[2:4]
        if month_assets_sum > 0:
            assets_result.append({"description": month_date, "sum": month_assets_sum})

        list_liabilities = await database.fetch_all(liabilities.select().where(and_(liabilities.c.owner_id == user_id,
                                                                                    liabilities.c.date_in <= date,
                                                                                    date <= liabilities.c.date_out)))
        month_liabilities = [dict(result) for result in list_liabilities]
        month_liabilities_sum = 0
        for month_asset in month_liabilities:
            month_liabilities_sum += month_asset['sum']
        month_date = str(date)[5:7] + "." + str(date)[2:4]
        if month_liabilities_sum > 0:
            liabilities_result.append({ "description": month_date, "sum": month_liabilities_sum})

        dt = this_month_begin - timedelta(seconds=1)
        if len(list_assets) == 0 and len(list_liabilities) == 0:
            break

    out.update({"assets": assets_result})
    out.update({"liabilities": liabilities_result})

    inflow_regular_result = dict()
    query = "SELECT date, description, sum FROM inflow WHERE owner_id = :owner_id AND description IN " \
            "(SELECT description FROM inflow_regular WHERE owner_id = :owner_id)"
    inflow_regular_result = await database.fetch_all(query=query, values={"owner_id": user_id})
    out.update({"inflow_regular": [dict(result) for result in inflow_regular_result]})

    outflow_regular_result = dict()
    query = "SELECT date, description, sum FROM outflow WHERE owner_id = :owner_id AND description IN " \
            "(SELECT description FROM outflow_regular WHERE owner_id = :owner_id)"
    outflow_regular_result = await database.fetch_all(query=query, values={"owner_id": user_id})
    out.update({"outflow_regular": [dict(result) for result in outflow_regular_result]})

    return out


async def get_export(user_id: int):
    wb = load_workbook('.' + os.sep + 'static' + os.sep + 'export' + os.sep + 'template.xlsx')
    bold = Font(bold=True)

    this_month_end = datetime.now()
    report = dict()

    # забираем категории пользователя
    db_user_categories = await get_user_categories(user_id)
    user_categories = {result['id']:result['category'] for result in db_user_categories['categories']}
    user_categories[None]='Без категории'

    while True:
        this_month_begin = datetime.strptime(f"{this_month_end.timetuple().tm_year}-{this_month_end.timetuple().tm_mon}"
                                             f"-01 00:00:00", "%Y-%m-%d %H:%M:%S")

        if this_month_end.timetuple().tm_mon < 10:
            this_month = f'0{this_month_end.timetuple().tm_mon}'
        else:
            this_month = f'{this_month_end.timetuple().tm_mon}'
        this_year_month = f"{this_month_end.timetuple().tm_year}-{this_month}"

        report[this_year_month] = dict()

        wb.create_sheet(this_year_month)
        sht = wb[this_year_month]
        sht.column_dimensions["A"].width = 32
        sht.column_dimensions["B"].width = 10
        sht.column_dimensions["C"].width = 8
        sht.column_dimensions["D"].width = 32
        sht.column_dimensions["E"].width = 12

        # экспорт доходов
        cashflow = 0
        inflows = await get_inflow_user(user_id, this_month_begin, this_month_end)
        cell = sht.cell(row=1, column=1)
        cell.value = 'Доходы'
        cell.font = bold
        inflow_row = 1
        inflow_sum = 0
        row = 2
        column = 1
        for month_inflows in inflows['inflow']:
            cell = sht.cell(row=row, column=column)
            cell.value = month_inflows['description']
            cell = sht.cell(row=row, column=column+1)
            cell.value = month_inflows['sum']
            inflow_sum += month_inflows['sum']
            cashflow += month_inflows['sum']
            row += 1

        # экспорт расходов
        row += 1
        outflows = await get_outflow_user(user_id, this_month_begin, this_month_end)
        cell = sht.cell(row=row, column=column)
        cell.value = 'Расходы'
        cell.font = bold
        outflow_row = row
        outflow_sum = 0
        row += 1
        for month_outflows in outflows['outflow']:
            cell = sht.cell(row=row, column=column)
            cell.value = month_outflows['description']
            cell = sht.cell(row=row, column=column+1)
            cell.value = month_outflows['sum']
            outflow_sum += month_outflows['sum']
            cashflow -= month_outflows['sum']
            row += 1

        # расчте денежного потока и общих сумм доходов и расходов
        row += 1
        cell = sht.cell(row=row, column=column)
        cell.value = 'Денежный поток'
        cell.font = bold
        cell = sht.cell(row=row, column=column+1)
        cell.value = cashflow
        cell.font = bold

        cell = sht.cell(row=inflow_row, column=column+1)
        cell.value = inflow_sum
        cell.font = bold
        cell = sht.cell(row=outflow_row, column=column+1)
        cell.value = outflow_sum
        cell.font = bold

        # экспорт активов
        capital = 0
        assets = await get_assets_user(user_id, this_month_begin+timedelta(days=15))
        cell = sht.cell(row=1, column=4)
        cell.value = 'Активы'
        cell.font = bold
        assets_row = 1
        assets_sum = 0
        row = 2
        column = 4
        for month_assets in assets['assets']:
            cell = sht.cell(row=row, column=column)
            cell.value = month_assets['description']
            cell = sht.cell(row=row, column=column+1)
            cell.value = month_assets['sum']
            assets_sum += month_assets['sum']
            capital += month_assets['sum']
            row += 1

        # экспорт пассивов
        row += 1
        liabilities = await get_liabilities_user(user_id, this_month_begin+timedelta(days=15))
        cell = sht.cell(row=row, column=column)
        cell.value = 'Пассивы'
        cell.font = bold
        liabilities_row = row
        liabilities_sum = 0
        row += 1
        for month_liabilities in liabilities['liabilities']:
            cell = sht.cell(row=row, column=column)
            cell.value = month_liabilities['description']
            cell = sht.cell(row=row, column=column+1)
            cell.value = month_liabilities['sum']
            liabilities_sum += month_liabilities['sum']
            capital -= month_liabilities['sum']
            row += 1

        # расчет капитала и общих сумм активов и пассивов
        row += 1
        cell = sht.cell(row=row, column=column)
        cell.value = 'Капитал'
        cell.font = bold
        cell = sht.cell(row=row, column=column+1)
        cell.value = capital
        cell.font = bold

        # суммы по категориям активов и пассивов
        row += 2
        cell = sht.cell(row=row, column=column)
        cell.value = 'По категориям'
        cell.font = bold

        asset_by_categories = await get_assets_by_categories(user_id, this_month_begin+timedelta(days=15), "assets")
        row += 1
        for month_asset_by_categories in asset_by_categories['categories']:
            cell = sht.cell(row=row, column=column)
            if month_asset_by_categories['category_id'] is not None:
                cell.value = user_categories[month_asset_by_categories['category_id']]
                report[this_year_month][user_categories[month_asset_by_categories['category_id']]] = \
                    month_asset_by_categories['sum']
            else:
                cell.value = "Активы без категории"
                report[this_year_month]["Активы без категории"] = month_asset_by_categories['sum']
            cell = sht.cell(row=row, column=column + 1)
            cell.value = month_asset_by_categories['sum']
            row += 1

        liabilities_by_categories = await get_assets_by_categories(user_id, this_month_begin+timedelta(days=15),
                                                                   "liabilities")
        for month_liabilities_by_categories in liabilities_by_categories['categories']:
            cell = sht.cell(row=row, column=column)
            if month_liabilities_by_categories['category_id'] is not None:
                cell.value = user_categories[month_liabilities_by_categories['category_id']]
                report[this_year_month][user_categories[month_liabilities_by_categories['category_id']]] = \
                    month_liabilities_by_categories['sum']
            else:
                cell.value = "Пассивы без категории"
                report[this_year_month]["Пассивы без категории"] = month_liabilities_by_categories['sum']
            cell = sht.cell(row=row, column=column + 1)
            cell.value = month_liabilities_by_categories['sum']
            row += 1

        cell = sht.cell(row=assets_row, column=column+1)
        cell.value = assets_sum
        cell.font = bold
        cell = sht.cell(row=liabilities_row, column=column+1)
        cell.value = liabilities_sum
        cell.font = bold

        # переход к следующему месяцу
        this_month_end = this_month_begin - timedelta(seconds=1)
        if inflow_sum + outflow_sum == 0:
            sht = wb.get_sheet_by_name(this_year_month)
            wb.remove_sheet(sht)
            report.pop(this_year_month)
            break
        else:
            report[this_year_month]['inflow'] = inflow_sum
            report[this_year_month]['outflow'] = outflow_sum
            report[this_year_month]['assets'] = assets_sum
            report[this_year_month]['liabilities'] = liabilities_sum

    sht = wb.get_sheet_by_name('clear')
    wb.remove_sheet(sht)

    wb.create_sheet('Свод', 0)
    sht = wb['Свод']

    for symbol in range(65, 80):
        sht.column_dimensions[chr(symbol)].width = 10

    cell = sht.cell(row=1, column=1)
    cell.value = 'Дата'
    cell.font = bold
    cell = sht.cell(row=1, column=2)
    cell.value = 'Доходы'
    cell.font = bold
    cell = sht.cell(row=1, column=3)
    cell.value = 'Расходы'
    cell.font = bold
    cell = sht.cell(row=1, column=4)
    cell.value = 'Cahflow'
    cell.font = bold
    cell = sht.cell(row=1, column=5)
    cell.value = 'Активы'
    cell.font = bold
    cell = sht.cell(row=1, column=6)
    cell.value = 'Пассивы'
    cell.font = bold
    cell = sht.cell(row=1, column=7)
    cell.value = 'Капитал'
    cell.font = bold

    column = 8
    for category in user_categories:
        if user_categories[category] != "Без категории":
            cell = sht.cell(row=1, column=column)
            cell.value = user_categories[category]
            cell.font = bold
            column += 1
        else:
            cell = sht.cell(row=1, column=column)
            cell.value = "Активы без категории"
            cell.font = bold
            column += 1
            cell = sht.cell(row=1, column=column)
            cell.value = "Пассивы без категории"
            cell.font = bold
            column += 1

    report = dict(sorted(report.items(), key=lambda x: x[0]))
    row = 1
    for month in report:
        row += 1
        cell = sht.cell(row=row, column=1)
        cell.value = month
        cell = sht.cell(row=row, column=2)
        cell.value = report[month]['inflow']
        cell = sht.cell(row=row, column=3)
        cell.value = report[month]['outflow']
        cell = sht.cell(row=row, column=4)
        cell.value = report[month]['inflow'] - report[month]['outflow']
        cell = sht.cell(row=row, column=5)
        cell.value = report[month]['assets']
        cell = sht.cell(row=row, column=6)
        cell.value = report[month]['liabilities']
        cell = sht.cell(row=row, column=7)
        cell.value = report[month]['assets'] - report[month]['liabilities']

        column = 8
        for category in user_categories:
            if user_categories[category] != "Без категории":
                if user_categories[category] in report[month]:
                    cell = sht.cell(row=row, column=column)
                    cell.value = report[month][user_categories[category]]
                else:
                    cell = sht.cell(row=row, column=column)
                    cell.value = 0
                column += 1
            else:
                if "Активы без категории" in report[month]:
                    cell = sht.cell(row=row, column=column)
                    cell.value = report[month]["Активы без категории"]
                else:
                    cell = sht.cell(row=row, column=column)
                    cell.value = 0
                column += 1
                if "Пассивы без категории" in report[month]:
                    cell = sht.cell(row=row, column=column)
                    cell.value = report[month]["Пассивы без категории"]
                else:
                    cell = sht.cell(row=row, column=column)
                    cell.value = 0
                column += 1

    wb.save('.' + os.sep + 'static' + os.sep + 'export' + os.sep + f'cashflow{user_id}.xlsx')


async def get_most_popular(user_id: int, date_in: datetime, date_out: datetime):
    result = dict()
    query = "SELECT description, COUNT(description) AS sum FROM outflow " \
            "WHERE owner_id = :owner_id AND description NOT IN " \
            "(SELECT description FROM outflow WHERE owner_id = :owner_id " \
            "AND :date_in < date AND date < :date_out) " \
            "AND description NOT IN (SELECT description FROM outflow_regular WHERE owner_id = :owner_id)" \
            "GROUP BY description " \
            "ORDER BY sum DESC LIMIT 3"

    list_most_popular = await database.fetch_all(query=query, values={"owner_id": user_id, "date_in": date_in,
                                                                      "date_out": date_out})

    result.update({"most_popular": [dict(result) for result in list_most_popular]})

    query = "SELECT DISTINCT description FROM inflow WHERE owner_id = :owner_id UNION " \
            "SELECT DISTINCT description FROM outflow WHERE owner_id = :owner_id"

    list_autocomplete = await database.fetch_all(query=query, values={"owner_id": user_id})

    result.update({"autocomplete": list_autocomplete})


    return result